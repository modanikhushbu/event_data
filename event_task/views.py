from bs4 import BeautifulSoup
from django.shortcuts import render
from django.http import HttpResponse
import requests
import itertools
import openpyxl
from events.models import Unintresting_url,Interesting_url

def Home(request):
    return render(request,'home.html')

# 'G:\Django-projects\event_task\sheet1.xlsx'
def group_list():
    # Give the location of the file
    path = 'G:\Django-projects\event_task\sheet1.xlsx'

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    group_list = []
    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 5)
        group_list.append(cell_obj.value)
    return(group_list)

def make_list(request):
    if request.method=='POST':
        url = request.POST['url']
        if 'insider' in url:
            l = dict()
            r = requests.get(url).text
            soup = BeautifulSoup(r,'html.parser')

            for a in soup.find_all(class_='event-card'):
                u = ((a.findChild())['href'])

                y = ((a.findChild()).findChild().text).rstrip('BUY')
                url = "https://insider.in" + u
                l[url]= y

            gl = group_list()
            li = []
            uli = []
            for key,val in list(itertools.islice(l.items(), 10)):
                if val in gl:
                    li.append(scrap_data(key,val))

                else:
                    full_url = "https://insider.in" + key
                    uli.append(full_url)

            uninteresting_url(uli)
            interesting_url(li)
            return HttpResponse('<br><br><br><center><h1>Data submitted</h1></center>')
        else:
            url_list = list()
            r = requests.get(url).text
            soup = BeautifulSoup(r, 'html.parser')

            for a in soup.find_all(class_='tribe-events-event-meta'):
                url = ((a.findChild())['href'])
                url_list.append(url)
            category = 'Yoga'
            gl = group_list()
            li = []
            uli = []
            url_list = url_list[:10]
            for i in url_list:
                if category in gl:
                    li.append(scrap_data(i, category))
                else:
                    uli.append(i)
            interesting_url(li)
            uninteresting_url(uli)
            return HttpResponse('<br><br><br><center><h1>Data submitted</h1></center>')

def scrap_data(full_url,val):
    if 'insider' in full_url:
        r = requests.get(full_url).text
        soup = BeautifulSoup(r, 'html.parser')
        title = soup.title.text
        for a in soup.find_all(class_='css-1h4eg37', limit=1):
            content = a.text
        return title, content, full_url, val

    else:
        r = requests.get(full_url).text
        soup = BeautifulSoup(r, 'html.parser')
        title = soup.title.text
        abc = [full_url,title,val ]
        for a in soup.find_all(class_=['date','time']):
            abc.append(((a.text).strip('\n')))
        return(abc)

def interesting_url(li):
    if li:
        if len(li[0])==4:
            for i in li:
                if Interesting_url.objects.filter(url =i[2]).exists():
                    pass
                else:
                    url = Interesting_url.objects.create(title=i[0], date_and_time=i[1], url=i[2],interested_group=i[3])

        else:
            for i in li:
                if Interesting_url.objects.filter(url =i[0]).exists():
                    pass
                else:
                    url = Interesting_url.objects.create(url = i[0],title = i[1],date_and_time =i[3]+i[4],interested_group = i[2])
                    url.save()

    else:
        print('No interesting event ur to save')

def uninteresting_url(uli):
    if uli:
        for u in uli:
            if Unintresting_url.objects.filter(url=u).exists():
                pass
            else:
                url = Unintresting_url.objects.create(url = u)
                url.save()

    else:
        print('No uninteresting url to save')






